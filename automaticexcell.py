from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, GradientFill, Alignment
import calendar
import time as _time
import math as _math
import sys


# Criar um novo workbook
#wb = load_workbook(r"C:\Users\Urubu\OneDrive\Documentos\GitHub\You-are-my-Sunshine\yha.xlsx")
wb = load_workbook(r"C:\Users\Alunos\Downloads\yha.xlsx")
WS = wb.active
WS.column_dimensions['C'].width = 35
WS.row_dimensions[8].height = 5
WS.row_dimensions[17].height = 5
linhazul_fill = PatternFill("solid", fgColor="041468")
medium = Side(border_style="medium")
dotted = Side(border_style="dotted")
thin = Side(border_style="thin")
WS['C8'].fill = linhazul_fill
WS['C17'].fill = linhazul_fill

#escritorio nomes
WS['C9'] = 'José António'
WS['C9'].font = Font(name="Calibri", b="true", size=18)
WS['C9'].alignment = Alignment(horizontal="right", vertical="center")
WS['C9'].border = Border(bottom=dotted)

WS['C10'] = 'Henrique Leal'
WS['C10'].font = Font(name="Calibri", b="true", size=18)
WS['C10'].alignment = Alignment(horizontal="right", vertical="center")
WS['C10'].border = Border(bottom=dotted)

WS['C11'] = 'Isabel Bastos'
WS['C11'].font = Font(name="Calibri", b="true", size=18)
WS['C11'].alignment = Alignment(horizontal="right", vertical="center")
WS['C11'].border = Border(bottom=dotted)

WS['C12'] = 'Tetiana Kalichenko'
WS['C12'].font = Font(name="Calibri", b="true", size=18)
WS['C12'].alignment = Alignment(horizontal="right", vertical="center")
WS['C12'].border = Border(bottom=dotted)

WS['C13'] = 'Helena Duarte'
WS['C13'].font = Font(name="Calibri", b="true", size=18)
WS['C13'].alignment = Alignment(horizontal="right", vertical="center")
WS['C13'].border = Border(bottom=dotted)

WS['C14'] = 'Andreia Pinto'
WS['C14'].font = Font(name="Calibri", b="true", size=18)
WS['C14'].alignment = Alignment(horizontal="right", vertical="center")
WS['C14'].border = Border(bottom=dotted)

WS['C15'] = 'André Pais Vieira'
WS['C15'].font = Font(name="Calibri", b="true", size=18)
WS['C15'].alignment = Alignment(horizontal="right", vertical="center")
WS['C15'].border = Border(bottom=dotted)

WS['C16'] = 'Rayanne dos Santos'
WS['C16'].font = Font(name="Calibri", b="true", size=18)
WS['C16'].alignment = Alignment(horizontal="right", vertical="center")
WS['C16'].border = Border(bottom=dotted)

#qualidade manha
WS['C18'] = 'Fábio Pereira'
WS['C18'].font = Font(name="Calibri", b="true", size=18)
WS['C18'].alignment = Alignment(horizontal="right", vertical="center")
WS['C18'].border = Border(bottom=dotted)
WS['C18'].fill = PatternFill("solid", fgColor="e8bcb4")

WS['C19'] = 'Miguel Pires'
WS['C19'].font = Font(name="Calibri", b="true", size=18)
WS['C19'].alignment = Alignment(horizontal="right", vertical="center")
WS['C19'].border = Border(bottom=dotted)
WS['C19'].fill = PatternFill("solid", fgColor="e8bcb4")

WS['C20'] = 'Cristina Tavares'
WS['C20'].font = Font(name="Calibri", b="true", size=18)
WS['C20'].alignment = Alignment(horizontal="right", vertical="center")
WS['C20'].border = Border(bottom=dotted)
WS['C20'].fill = PatternFill("solid", fgColor="e8bcb4")

WS['C21'] = 'Bianca Gonçalves'
WS['C21'].font = Font(name="Calibri", b="true", size=18)
WS['C21'].alignment = Alignment(horizontal="right", vertical="center")
WS['C21'].border = Border(bottom=dotted)
WS['C21'].fill = PatternFill("solid", fgColor="e8bcb4")

WS['C22'] = 'Liudmyla Murmryk'
WS['C22'].font = Font(name="Calibri", b="true", size=18)
WS['C22'].alignment = Alignment(horizontal="right", vertical="center")
WS['C22'].border = Border(bottom=dotted)
WS['C22'].fill = PatternFill("solid", fgColor="e8bcb4")

WS['C23'] = 'Vânia Correia'
WS['C23'].font = Font(name="Calibri", b="true", size=18)
WS['C23'].alignment = Alignment(horizontal="right", vertical="center")
WS['C23'].border = Border(bottom=dotted)
WS['C23'].fill = PatternFill("solid", fgColor="e8bcb4")

# Mapeamento dos dias da semana em inglês para português de Portugal
dias_semana_pt = {
    'Mon': 'Seg',
    'Tue': 'Ter',
    'Wed': 'Qua',
    'Thu': 'Qui',
    'Fri': 'Sex',
    'Sat': 'Sáb',
    'Sun': 'Dom'
}

#Dias da semana e do mes
# Definir cores para sábado e domingo
fimdesemana_fill = PatternFill("solid", fgColor="fffffccc")

# Definir o mês e o ano
ano = 2024
mes = 5

# Obter o número de dias no mês
num_dias = calendar.monthrange(ano, mes)[1]

# Definir datas iniciais
start_date = f"{ano}-{mes}-01"

num_pessoas_escritorio = 8
num_pessoas_qualidade = 6

folga_fill = PatternFill("solid", fgColor="ff9ccc")

# Percorrer os dias do mês
for col in range(4, num_dias + 4):
    current_date = start_date.split("-")[-1]
    # Obter o dia da semana em inglês
    dia_semana_en = calendar.day_abbr[calendar.weekday(ano, mes, int(current_date))]
    # Traduzir para português de Portugal
    dia_semana_pt = dias_semana_pt[dia_semana_en]
    
    if dia_semana_en == 'Sat':  # Sábado
        WS[get_column_letter(col) + '3'].fill = fimdesemana_fill
        WS[get_column_letter(col) + '4'].fill = fimdesemana_fill
    elif dia_semana_en == 'Sun':  # Domingo
        WS[get_column_letter(col) + '3'].fill = fimdesemana_fill
        WS[get_column_letter(col) + '4'].fill = fimdesemana_fill

    #escritorio
    for lin in range(9, num_pessoas_escritorio + 9):
        WS.row_dimensions[lin].height = 25
        WS[get_column_letter(col) + str(lin)].alignment = Alignment(horizontal="center", vertical="center")
        WS[get_column_letter(col) + str(lin)].font = Font(name="Calibri", size=18)
        WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=dotted, left=dotted, bottom=dotted)
        if dia_semana_en == 'Sat':  # Sábado
            WS[get_column_letter(col) + str(lin)] = 'D'
            WS[get_column_letter(col) + str(lin)].fill = folga_fill
        elif dia_semana_en == 'Sun':  # Domingo
            WS[get_column_letter(col) + str(lin)] = 'D'
            WS[get_column_letter(col) + str(lin)].fill = folga_fill
        elif dia_semana_en == 'Mon' or 'Tue' or 'Wed' or 'Thu' or 'Fri':
            WS[get_column_letter(col) + str(lin)] = 'E'
        
        if lin == 9:
            WS[get_column_letter(col) + str(lin)].border = Border(top=medium, right=dotted, left=dotted, bottom=dotted)

        if col == 4:
            WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=dotted, left=medium, bottom=dotted)
            WS[get_column_letter(col) + '9'].border = Border(top=medium, right=dotted, left=medium, bottom=dotted)
        elif col == 33:
            WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=medium, left=dotted, bottom=dotted)
            WS[get_column_letter(col) + '9'].border = Border(top=medium, right=medium, left=dotted, bottom=dotted)
        elif col == 34:
            WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=medium, left=dotted, bottom=dotted)
            WS[get_column_letter(col) + '9'].border = Border(top=medium, right=medium, left=dotted, bottom=dotted)
            WS[get_column_letter(col-1) + str(lin)].border = Border(top=dotted, right=dotted, left=dotted, bottom=dotted)
            WS[get_column_letter(col-1) + '9'].border = Border(top=medium, right=dotted, left=dotted, bottom=dotted)

    #qualidade
    #manha
    for lin in range(18, num_pessoas_qualidade + 18):
        WS.row_dimensions[lin].height = 25
        WS[get_column_letter(col) + str(lin)].alignment = Alignment(horizontal="center", vertical="center")
        WS[get_column_letter(col) + str(lin)].font = Font(name="Calibri", size=18)
        WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=dotted, left=dotted, bottom=dotted)
        if dia_semana_en == 'Sat':  # Sábado
            WS[get_column_letter(col) + str(lin)] = 'D'
            WS[get_column_letter(col) + str(lin)].fill = folga_fill
        elif dia_semana_en == 'Sun':  # Domingo
            WS[get_column_letter(col) + str(lin)] = 'D'
            WS[get_column_letter(col) + str(lin)].fill = folga_fill
        elif dia_semana_en == 'Mon' or 'Tue' or 'Wed' or 'Thu' or 'Fri' and lin <= 19:
            if lin <= 19:
                WS[get_column_letter(col) + str(lin)] = 'MQ'
            elif lin == 20:
                WS[get_column_letter(col) + str(lin)] = 'TQ'
            elif lin == 21:
                WS[get_column_letter(col) + str(lin)] = 'NQ'
            elif lin == 22:
                WS[get_column_letter(col) + str(lin)] = 'TQ'
            elif lin == 23:
                WS[get_column_letter(col) + str(lin)] = 'MQ'
        
        if col == 4:
            WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=dotted, left=medium, bottom=dotted)
        elif col == 33:
            WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=medium, left=dotted, bottom=dotted)
        elif col == 34:
            WS[get_column_letter(col) + str(lin)].border = Border(top=dotted, right=medium, left=dotted, bottom=dotted)
            WS[get_column_letter(col-1) + str(lin)].border = Border(top=dotted, right=dotted, left=dotted, bottom=dotted)

    WS.column_dimensions[get_column_letter(col)].width = 7

    WS[get_column_letter(col) + '3'] = dia_semana_pt
    WS[get_column_letter(col) + '3'].font = Font(name="Calibri", b="true", size=18)
    WS[get_column_letter(col) + '3'].alignment = Alignment(horizontal="center", vertical="center")
    WS[get_column_letter(col) + '3'].border = Border(top=dotted, right=dotted, left=dotted, bottom=dotted)

    WS[get_column_letter(col) + '4'] = current_date
    WS[get_column_letter(col) + '4'].font = Font(name="Calibri", b="true", size=18)
    WS[get_column_letter(col) + '4'].alignment = Alignment(horizontal="center", vertical="center")
    WS[get_column_letter(col) + '4'].border = Border(right=dotted, left=dotted, bottom=medium)

    WS[get_column_letter(col) + '8'].fill = linhazul_fill
    WS[get_column_letter(col) + '17'].fill = linhazul_fill

    if col == 4:
        WS[get_column_letter(col) + '3'].border = Border(top=dotted, right=dotted, left=medium, bottom=dotted)
        
        WS[get_column_letter(col) + '4'] = '1'
        WS[get_column_letter(col) + '4'].border = Border(right=dotted, left=medium, bottom=medium)

        WS[get_column_letter(col) + '8'].border = Border(top=dotted, right=dotted, left=medium, bottom=thin)
        WS[get_column_letter(col) + '17'].border = Border(top=dotted, right=dotted, left=medium, bottom=thin)
    elif col == 33:
        WS[get_column_letter(col) + '3'].border = Border(top=dotted, right=medium, left=dotted, bottom=dotted)

        WS[get_column_letter(col) + '4'].border = Border(right=medium, left=dotted, bottom=medium)

        WS[get_column_letter(col) + '8'].border = Border(top=dotted, right=medium, left=dotted, bottom=thin)
        WS[get_column_letter(col) + '17'].border = Border(top=dotted, right=medium, left=dotted, bottom=thin)
    elif col == 34:
        WS[get_column_letter(col) + '3'].border = Border(top=dotted, right=medium, left=dotted, bottom=dotted)
        WS[get_column_letter(col-1) + '3'].border = Border(top=dotted, right=dotted, left=dotted, bottom=dotted)

        WS[get_column_letter(col) + '4'].border = Border(right=medium, left=dotted, bottom=medium)
        WS[get_column_letter(col-1) + '4'].border = Border(right=dotted, left=dotted, bottom=medium)

        WS[get_column_letter(col) + '8'].border = Border(top=dotted, right=medium, left=dotted, bottom=thin)
        WS[get_column_letter(col) + '17'].border = Border(top=dotted, right=medium, left=dotted, bottom=thin)
        WS[get_column_letter(col-1) + '8'].border = Border(top=dotted, right=dotted, left=dotted, bottom=thin)
        WS[get_column_letter(col-1) + '17'].border = Border(top=dotted, right=dotted, left=dotted, bottom=thin)

    start_date = f"{ano}-{mes}-{int(current_date) + 1}"

#mês
# Mapeamento dos meses em inglês para português de Portugal
meses_pt = {
    'January': 'Janeiro',
    'February': 'Fevereiro',
    'March': 'Março',
    'April': 'Abril',
    'May': 'Maio',
    'June': 'Junho',
    'July': 'Julho',
    'August': 'Agosto',
    'September': 'Setembro',
    'October': 'Outubro',
    'November': 'Novembro',
    'December': 'Dezembro'
}

# Definir o nome do mês em inglês
nome_mes_en = calendar.month_name[mes]

# Traduzir para português de Portugal
nome_mes_pt = meses_pt[nome_mes_en]

if num_dias == 30:
    WS.merge_cells('D2:AG2')

if num_dias == 31:
    WS.merge_cells('D2:AH2')
# Obter a célula D2
d2 = WS['D2']
# Atribuir o nome do mês em português de Portugal à célula D2
WS['D2'] = nome_mes_pt.capitalize()  # Capitalize para a primeira letra em maiúscula
# Ajustar alinhamento, fonte, preenchimento e borda da célula D2
d2.alignment = Alignment(horizontal="center", vertical="center")
d2.font = Font(name="Calibri", size=48, b="true")
d2.fill = PatternFill("solid", fgColor="fffffccc")
d2.border = Border(top=medium, left=medium, right=medium)

for col in range(4, num_dias + 4):
    WS[get_column_letter(col) + '2'].border = Border(top=medium, left=medium, right=medium)

#total manha
WS['C5'] = "Total Manhã"
WS['C5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['C5'].font = Font(name="Calibri", b="true", size=18)
WS['C5'].alignment = Alignment(horizontal="right", vertical="center")
WS['C5'].border = Border(top=medium, right=medium)

WS['D5'] = '0'
WS['D5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['D5'].font = Font(name="Calibri", b="true", size=18)
WS['D5'].alignment = Alignment(horizontal="center", vertical="center")
WS['D5'].border = Border(top=medium, bottom=medium)

WS['E5'] = '0'
WS['E5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['E5'].font = Font(name="Calibri", b="true", size=18)
WS['E5'].alignment = Alignment(horizontal="center", vertical="center")
WS['E5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['F5'] = '0'
WS['F5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['F5'].font = Font(name="Calibri", b="true", size=18)
WS['F5'].alignment = Alignment(horizontal="center", vertical="center")
WS['F5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['G5'] = '0'
WS['G5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['G5'].font = Font(name="Calibri", b="true", size=18)
WS['G5'].alignment = Alignment(horizontal="center", vertical="center")
WS['G5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['H5'] = '0'
WS['H5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['H5'].font = Font(name="Calibri", b="true", size=18)
WS['H5'].alignment = Alignment(horizontal="center", vertical="center")
WS['H5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['I5'] = '0'
WS['I5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['I5'].font = Font(name="Calibri", b="true", size=18)
WS['I5'].alignment = Alignment(horizontal="center", vertical="center")
WS['I5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['J5'] = '0'
WS['J5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['J5'].font = Font(name="Calibri", b="true", size=18)
WS['J5'].alignment = Alignment(horizontal="center", vertical="center")
WS['J5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['K5'] = '0'
WS['K5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['K5'].font = Font(name="Calibri", b="true", size=18)
WS['K5'].alignment = Alignment(horizontal="center", vertical="center")
WS['K5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['L5'] = '0'
WS['L5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['L5'].font = Font(name="Calibri", b="true", size=18)
WS['L5'].alignment = Alignment(horizontal="center", vertical="center")
WS['L5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['M5'] = '0'
WS['M5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['M5'].font = Font(name="Calibri", b="true", size=18)
WS['M5'].alignment = Alignment(horizontal="center", vertical="center")
WS['M5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['N5'] = '0'
WS['N5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['N5'].font = Font(name="Calibri", b="true", size=18)
WS['N5'].alignment = Alignment(horizontal="center", vertical="center")
WS['N5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['O5'] = '0'
WS['O5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['O5'].font = Font(name="Calibri", b="true", size=18)
WS['O5'].alignment = Alignment(horizontal="center", vertical="center")
WS['O5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['P5'] = '0'
WS['P5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['P5'].font = Font(name="Calibri", b="true", size=18)
WS['P5'].alignment = Alignment(horizontal="center", vertical="center")
WS['P5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['Q5'] = '0'
WS['Q5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['Q5'].font = Font(name="Calibri", b="true", size=18)
WS['Q5'].alignment = Alignment(horizontal="center", vertical="center")
WS['Q5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['R5'] = '0'
WS['R5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['R5'].font = Font(name="Calibri", b="true", size=18)
WS['R5'].alignment = Alignment(horizontal="center", vertical="center")
WS['R5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['S5'] = '0'
WS['S5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['S5'].font = Font(name="Calibri", b="true", size=18)
WS['S5'].alignment = Alignment(horizontal="center", vertical="center")
WS['S5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['T5'] = '0'
WS['T5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['T5'].font = Font(name="Calibri", b="true", size=18)
WS['T5'].alignment = Alignment(horizontal="center", vertical="center")
WS['T5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['U5'] = '0'
WS['U5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['U5'].font = Font(name="Calibri", b="true", size=18)
WS['U5'].alignment = Alignment(horizontal="center", vertical="center")
WS['U5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['V5'] = '0'
WS['V5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['V5'].font = Font(name="Calibri", b="true", size=18)
WS['V5'].alignment = Alignment(horizontal="center", vertical="center")
WS['V5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['W5'] = '0'
WS['W5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['W5'].font = Font(name="Calibri", b="true", size=18)
WS['W5'].alignment = Alignment(horizontal="center", vertical="center")
WS['W5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['X5'] = '0'
WS['X5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['X5'].font = Font(name="Calibri", b="true", size=18)
WS['X5'].alignment = Alignment(horizontal="center", vertical="center")
WS['X5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['Y5'] = '0'
WS['Y5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['Y5'].font = Font(name="Calibri", b="true", size=18)
WS['Y5'].alignment = Alignment(horizontal="center", vertical="center")
WS['Y5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['Z5'] = '0'
WS['Z5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['Z5'].font = Font(name="Calibri", b="true", size=18)
WS['Z5'].alignment = Alignment(horizontal="center", vertical="center")
WS['Z5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AA5'] = '0'
WS['AA5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['AA5'].font = Font(name="Calibri", b="true", size=18)
WS['AA5'].alignment = Alignment(horizontal="center", vertical="center")
WS['AA5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AB5'] = '0'
WS['AB5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['AB5'].font = Font(name="Calibri", b="true", size=18)
WS['AB5'].alignment = Alignment(horizontal="center", vertical="center")
WS['AB5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AC5'] = '0'
WS['AC5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['AC5'].font = Font(name="Calibri", b="true", size=18)
WS['AC5'].alignment = Alignment(horizontal="center", vertical="center")
WS['AC5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AD5'] = '0'
WS['AD5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['AD5'].font = Font(name="Calibri", b="true", size=18)
WS['AD5'].alignment = Alignment(horizontal="center", vertical="center")
WS['AD5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AE5'] = '0'
WS['AE5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['AE5'].font = Font(name="Calibri", b="true", size=18)
WS['AE5'].alignment = Alignment(horizontal="center", vertical="center")
WS['AE5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AF5'] = '0'
WS['AF5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['AF5'].font = Font(name="Calibri", b="true", size=18)
WS['AF5'].alignment = Alignment(horizontal="center", vertical="center")
WS['AF5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AG5'] = '0'
WS['AG5'].fill = PatternFill("solid", fgColor="a0fc9c")
WS['AG5'].font = Font(name="Calibri", b="true", size=18)
WS['AG5'].alignment = Alignment(horizontal="center", vertical="center")
WS['AG5'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

if num_dias == 30:
    WS['AG5'].border = Border(top=medium, bottom=dotted, right=medium, left=dotted)

if num_dias == 31:
    WS['AH5'] = '0'
    WS['AH5'].fill = PatternFill("solid", fgColor="a0fc9c")
    WS['AH5'].font = Font(name="Calibri", b="true", size=18)
    WS['AH5'].alignment = Alignment(horizontal="center", vertical="center")
    WS['AH5'].border = Border(top=medium, bottom=medium, right=medium, left=dotted)

#total tarde
WS['C6'] = "Total Tarde"
WS['C6'].fill = PatternFill("solid", fgColor="08b454")
WS['C6'].font = Font(name="Calibri", b="true", size=18)
WS['C6'].alignment = Alignment(horizontal="right", vertical="center")
WS['C6'].border = Border(top=medium, right=medium)

WS['D6'] = '0'
WS['D6'].fill = PatternFill("solid", fgColor="08b454")
WS['D6'].font = Font(name="Calibri", b="true", size=18)
WS['D6'].alignment = Alignment(horizontal="center", vertical="center")
WS['D6'].border = Border(top=medium, bottom=medium)

WS['E6'] = '0'
WS['E6'].fill = PatternFill("solid", fgColor="08b454")
WS['E6'].font = Font(name="Calibri", b="true", size=18)
WS['E6'].alignment = Alignment(horizontal="center", vertical="center")
WS['E6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['F6'] = '0'
WS['F6'].fill = PatternFill("solid", fgColor="08b454")
WS['F6'].font = Font(name="Calibri", b="true", size=18)
WS['F6'].alignment = Alignment(horizontal="center", vertical="center")
WS['F6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['G6'] = '0'
WS['G6'].fill = PatternFill("solid", fgColor="08b454")
WS['G6'].font = Font(name="Calibri", b="true", size=18)
WS['G6'].alignment = Alignment(horizontal="center", vertical="center")
WS['G6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['H6'] = '0'
WS['H6'].fill = PatternFill("solid", fgColor="08b454")
WS['H6'].font = Font(name="Calibri", b="true", size=18)
WS['H6'].alignment = Alignment(horizontal="center", vertical="center")
WS['H6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['I6'] = '0'
WS['I6'].fill = PatternFill("solid", fgColor="08b454")
WS['I6'].font = Font(name="Calibri", b="true", size=18)
WS['I6'].alignment = Alignment(horizontal="center", vertical="center")
WS['I6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['J6'] = '0'
WS['J6'].fill = PatternFill("solid", fgColor="08b454")
WS['J6'].font = Font(name="Calibri", b="true", size=18)
WS['J6'].alignment = Alignment(horizontal="center", vertical="center")
WS['J6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['K6'] = '0'
WS['K6'].fill = PatternFill("solid", fgColor="08b454")
WS['K6'].font = Font(name="Calibri", b="true", size=18)
WS['K6'].alignment = Alignment(horizontal="center", vertical="center")
WS['K6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['L6'] = '0'
WS['L6'].fill = PatternFill("solid", fgColor="08b454")
WS['L6'].font = Font(name="Calibri", b="true", size=18)
WS['L6'].alignment = Alignment(horizontal="center", vertical="center")
WS['L6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['M6'] = '0'
WS['M6'].fill = PatternFill("solid", fgColor="08b454")
WS['M6'].font = Font(name="Calibri", b="true", size=18)
WS['M6'].alignment = Alignment(horizontal="center", vertical="center")
WS['M6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['N6'] = '0'
WS['N6'].fill = PatternFill("solid", fgColor="08b454")
WS['N6'].font = Font(name="Calibri", b="true", size=18)
WS['N6'].alignment = Alignment(horizontal="center", vertical="center")
WS['N6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['O6'] = '0'
WS['O6'].fill = PatternFill("solid", fgColor="08b454")
WS['O6'].font = Font(name="Calibri", b="true", size=18)
WS['O6'].alignment = Alignment(horizontal="center", vertical="center")
WS['O6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['P6'] = '0'
WS['P6'].fill = PatternFill("solid", fgColor="08b454")
WS['P6'].font = Font(name="Calibri", b="true", size=18)
WS['P6'].alignment = Alignment(horizontal="center", vertical="center")
WS['P6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['Q6'] = '0'
WS['Q6'].fill = PatternFill("solid", fgColor="08b454")
WS['Q6'].font = Font(name="Calibri", b="true", size=18)
WS['Q6'].alignment = Alignment(horizontal="center", vertical="center")
WS['Q6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['R6'] = '0'
WS['R6'].fill = PatternFill("solid", fgColor="08b454")
WS['R6'].font = Font(name="Calibri", b="true", size=18)
WS['R6'].alignment = Alignment(horizontal="center", vertical="center")
WS['R6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['S6'] = '0'
WS['S6'].fill = PatternFill("solid", fgColor="08b454")
WS['S6'].font = Font(name="Calibri", b="true", size=18)
WS['S6'].alignment = Alignment(horizontal="center", vertical="center")
WS['S6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['T6'] = '0'
WS['T6'].fill = PatternFill("solid", fgColor="08b454")
WS['T6'].font = Font(name="Calibri", b="true", size=18)
WS['T6'].alignment = Alignment(horizontal="center", vertical="center")
WS['T6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['U6'] = '0'
WS['U6'].fill = PatternFill("solid", fgColor="08b454")
WS['U6'].font = Font(name="Calibri", b="true", size=18)
WS['U6'].alignment = Alignment(horizontal="center", vertical="center")
WS['U6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['V6'] = '0'
WS['V6'].fill = PatternFill("solid", fgColor="08b454")
WS['V6'].font = Font(name="Calibri", b="true", size=18)
WS['V6'].alignment = Alignment(horizontal="center", vertical="center")
WS['V6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['W6'] = '0'
WS['W6'].fill = PatternFill("solid", fgColor="08b454")
WS['W6'].font = Font(name="Calibri", b="true", size=18)
WS['W6'].alignment = Alignment(horizontal="center", vertical="center")
WS['W6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['X6'] = '0'
WS['X6'].fill = PatternFill("solid", fgColor="08b454")
WS['X6'].font = Font(name="Calibri", b="true", size=18)
WS['X6'].alignment = Alignment(horizontal="center", vertical="center")
WS['X6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['Y6'] = '0'
WS['Y6'].fill = PatternFill("solid", fgColor="08b454")
WS['Y6'].font = Font(name="Calibri", b="true", size=18)
WS['Y6'].alignment = Alignment(horizontal="center", vertical="center")
WS['Y6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['Z6'] = '0'
WS['Z6'].fill = PatternFill("solid", fgColor="08b454")
WS['Z6'].font = Font(name="Calibri", b="true", size=18)
WS['Z6'].alignment = Alignment(horizontal="center", vertical="center")
WS['Z6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AA6'] = '0'
WS['AA6'].fill = PatternFill("solid", fgColor="08b454")
WS['AA6'].font = Font(name="Calibri", b="true", size=18)
WS['AA6'].alignment = Alignment(horizontal="center", vertical="center")
WS['AA6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AB6'] = '0'
WS['AB6'].fill = PatternFill("solid", fgColor="08b454")
WS['AB6'].font = Font(name="Calibri", b="true", size=18)
WS['AB6'].alignment = Alignment(horizontal="center", vertical="center")
WS['AB6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AC6'] = '0'
WS['AC6'].fill = PatternFill("solid", fgColor="08b454")
WS['AC6'].font = Font(name="Calibri", b="true", size=18)
WS['AC6'].alignment = Alignment(horizontal="center", vertical="center")
WS['AC6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AD6'] = '0'
WS['AD6'].fill = PatternFill("solid", fgColor="08b454")
WS['AD6'].font = Font(name="Calibri", b="true", size=18)
WS['AD6'].alignment = Alignment(horizontal="center", vertical="center")
WS['AD6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AE6'] = '0'
WS['AE6'].fill = PatternFill("solid", fgColor="08b454")
WS['AE6'].font = Font(name="Calibri", b="true", size=18)
WS['AE6'].alignment = Alignment(horizontal="center", vertical="center")
WS['AE6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AF6'] = '0'
WS['AF6'].fill = PatternFill("solid", fgColor="08b454")
WS['AF6'].font = Font(name="Calibri", b="true", size=18)
WS['AF6'].alignment = Alignment(horizontal="center", vertical="center")
WS['AF6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

WS['AG6'] = '0'
WS['AG6'].fill = PatternFill("solid", fgColor="08b454")
WS['AG6'].font = Font(name="Calibri", b="true", size=18)
WS['AG6'].alignment = Alignment(horizontal="center", vertical="center")
WS['AG6'].border = Border(top=medium, bottom=medium, right=dotted, left=dotted)

if num_dias == 30:
    WS['AG6'].border = Border(top=medium, bottom=dotted, right=medium, left=dotted)

if num_dias == 31:
    WS['AH6'] = '0'
    WS['AH6'].fill = PatternFill("solid", fgColor="08b454")
    WS['AH6'].font = Font(name="Calibri", b="true", size=18)
    WS['AH6'].alignment = Alignment(horizontal="center", vertical="center")
    WS['AH6'].border = Border(top=medium, bottom=medium, right=medium, left=dotted)

#total noite
WS['C7'] = "Total Noite"
WS['C7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['C7'].font = Font(name="Calibri", b="true", size=18)
WS['C7'].alignment = Alignment(horizontal="right", vertical="center")
WS['C7'].border = Border(top=medium, right=medium, bottom=dotted)

WS['D7'] = '0'
WS['D7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['D7'].font = Font(name="Calibri", b="true", size=18)
WS['D7'].alignment = Alignment(horizontal="center", vertical="center")
WS['D7'].border = Border(top=medium, bottom=dotted)

WS['E7'] = '0'
WS['E7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['E7'].font = Font(name="Calibri", b="true", size=18)
WS['E7'].alignment = Alignment(horizontal="center", vertical="center")
WS['E7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['F7'] = '0'
WS['F7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['F7'].font = Font(name="Calibri", b="true", size=18)
WS['F7'].alignment = Alignment(horizontal="center", vertical="center")
WS['F7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['G7'] = '0'
WS['G7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['G7'].font = Font(name="Calibri", b="true", size=18)
WS['G7'].alignment = Alignment(horizontal="center", vertical="center")
WS['G7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['H7'] = '0'
WS['H7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['H7'].font = Font(name="Calibri", b="true", size=18)
WS['H7'].alignment = Alignment(horizontal="center", vertical="center")
WS['H7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['I7'] = '0'
WS['I7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['I7'].font = Font(name="Calibri", b="true", size=18)
WS['I7'].alignment = Alignment(horizontal="center", vertical="center")
WS['I7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['J7'] = '0'
WS['J7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['J7'].font = Font(name="Calibri", b="true", size=18)
WS['J7'].alignment = Alignment(horizontal="center", vertical="center")
WS['J7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['K7'] = '0'
WS['K7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['K7'].font = Font(name="Calibri", b="true", size=18)
WS['K7'].alignment = Alignment(horizontal="center", vertical="center")
WS['K7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['L7'] = '0'
WS['L7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['L7'].font = Font(name="Calibri", b="true", size=18)
WS['L7'].alignment = Alignment(horizontal="center", vertical="center")
WS['L7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['M7'] = '0'
WS['M7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['M7'].font = Font(name="Calibri", b="true", size=18)
WS['M7'].alignment = Alignment(horizontal="center", vertical="center")
WS['M7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['N7'] = '0'
WS['N7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['N7'].font = Font(name="Calibri", b="true", size=18)
WS['N7'].alignment = Alignment(horizontal="center", vertical="center")
WS['N7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['O7'] = '0'
WS['O7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['O7'].font = Font(name="Calibri", b="true", size=18)
WS['O7'].alignment = Alignment(horizontal="center", vertical="center")
WS['O7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['P7'] = '0'
WS['P7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['P7'].font = Font(name="Calibri", b="true", size=18)
WS['P7'].alignment = Alignment(horizontal="center", vertical="center")
WS['P7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['Q7'] = '0'
WS['Q7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['Q7'].font = Font(name="Calibri", b="true", size=18)
WS['Q7'].alignment = Alignment(horizontal="center", vertical="center")
WS['Q7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['R7'] = '0'
WS['R7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['R7'].font = Font(name="Calibri", b="true", size=18)
WS['R7'].alignment = Alignment(horizontal="center", vertical="center")
WS['R7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['S7'] = '0'
WS['S7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['S7'].font = Font(name="Calibri", b="true", size=18)
WS['S7'].alignment = Alignment(horizontal="center", vertical="center")
WS['S7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['T7'] = '0'
WS['T7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['T7'].font = Font(name="Calibri", b="true", size=18)
WS['T7'].alignment = Alignment(horizontal="center", vertical="center")
WS['T7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['U7'] = '0'
WS['U7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['U7'].font = Font(name="Calibri", b="true", size=18)
WS['U7'].alignment = Alignment(horizontal="center", vertical="center")
WS['U7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['V7'] = '0'
WS['V7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['V7'].font = Font(name="Calibri", b="true", size=18)
WS['V7'].alignment = Alignment(horizontal="center", vertical="center")
WS['V7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['W7'] = '0'
WS['W7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['W7'].font = Font(name="Calibri", b="true", size=18)
WS['W7'].alignment = Alignment(horizontal="center", vertical="center")
WS['W7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['X7'] = '0'
WS['X7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['X7'].font = Font(name="Calibri", b="true", size=18)
WS['X7'].alignment = Alignment(horizontal="center", vertical="center")
WS['X7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['Y7'] = '0'
WS['Y7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['Y7'].font = Font(name="Calibri", b="true", size=18)
WS['Y7'].alignment = Alignment(horizontal="center", vertical="center")
WS['Y7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['Z7'] = '0'
WS['Z7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['Z7'].font = Font(name="Calibri", b="true", size=18)
WS['Z7'].alignment = Alignment(horizontal="center", vertical="center")
WS['Z7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['AA7'] = '0'
WS['AA7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['AA7'].font = Font(name="Calibri", b="true", size=18)
WS['AA7'].alignment = Alignment(horizontal="center", vertical="center")
WS['AA7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['AB7'] = '0'
WS['AB7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['AB7'].font = Font(name="Calibri", b="true", size=18)
WS['AB7'].alignment = Alignment(horizontal="center", vertical="center")
WS['AB7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['AC7'] = '0'
WS['AC7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['AC7'].font = Font(name="Calibri", b="true", size=18)
WS['AC7'].alignment = Alignment(horizontal="center", vertical="center")
WS['AC7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['AD7'] = '0'
WS['AD7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['AD7'].font = Font(name="Calibri", b="true", size=18)
WS['AD7'].alignment = Alignment(horizontal="center", vertical="center")
WS['AD7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['AE7'] = '0'
WS['AE7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['AE7'].font = Font(name="Calibri", b="true", size=18)
WS['AE7'].alignment = Alignment(horizontal="center", vertical="center")
WS['AE7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['AF7'] = '0'
WS['AF7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['AF7'].font = Font(name="Calibri", b="true", size=18)
WS['AF7'].alignment = Alignment(horizontal="center", vertical="center")
WS['AF7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

WS['AG7'] = '0'
WS['AG7'].fill = PatternFill("solid", fgColor="c8d49c")
WS['AG7'].font = Font(name="Calibri", b="true", size=18)
WS['AG7'].alignment = Alignment(horizontal="center", vertical="center")
WS['AG7'].border = Border(top=medium, bottom=dotted, right=dotted, left=dotted)

if num_dias == 30:
    WS['AG7'].border = Border(top=medium, bottom=dotted, right=medium, left=dotted)

if num_dias == 31:
    WS['AH7'] = '0' 
    WS['AH7'].fill = PatternFill("solid", fgColor="c8d49c")
    WS['AH7'].font = Font(name="Calibri", b="true", size=18)
    WS['AH7'].alignment = Alignment(horizontal="center", vertical="center")
    WS['AH7'].border = Border(top=medium, bottom=dotted, right=medium, left=dotted)

#escritorio


wb.save(r"C:\Users\Alunos\Downloads\yha.xlsx")
#wb.save(r"C:\Users\Urubu\OneDrive\Documentos\GitHub\You-are-my-Sunshine\yha.xlsx")
